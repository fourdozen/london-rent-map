import openpyxl as xl
import os

# Take Standard excel file from Mayor's office data release and keep only relevant data
class ExcelCleaner():
    
    def __init__(self, filename):
        self.filename = filename
        self.filepath = self._data_filepath(self.filename)
        self.workbook  = self._get_workbook()

    @staticmethod
    def _data_filepath(filename):
        cwd = os.getcwd()
        filepath = os.path.join(cwd, "data/excel", filename)
        return filepath
    
    def _get_workbook(self):
        return xl.load_workbook(self.filepath, data_only = True)
    
    def _delete_other_worksheets (self, sheet_name = '3') -> xl.worksheet.worksheet.Worksheet:
        all_sheets = self.workbook.sheetnames
        for sheet in all_sheets:
            if sheet != sheet_name:
                self.workbook.remove(self.workbook[sheet])
        return self.workbook[sheet_name]
    
    @staticmethod
    def _clean_rows(worksheet):
        worksheet.delete_rows(1, 2)
        return worksheet
    
    @staticmethod
    def _remove_formatting(worksheet):
        for row in worksheet.iter_rows():
            for cell in row:
                cell.style = 'Normal'
        return worksheet

    def _save_clean_workbook(self):
        new_filename = "clean_" + self.filename
        new_filepath = self._data_filepath(new_filename)
        self.workbook.save(new_filepath)

    def clean(self):
        _ = self._get_workbook()
        ws = self._delete_other_worksheets()
        clean_ws = self._clean_rows(ws)
        clean_ws = self._remove_formatting(clean_ws)
        self._save_clean_workbook()


if __name__ == "__main__":
    cleaner = ExcelCleaner("londonrentalstatsaccessibleq32024.xlsx").clean()